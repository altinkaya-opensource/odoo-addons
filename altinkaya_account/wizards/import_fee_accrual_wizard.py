import base64
from datetime import date as dt_date
from datetime import datetime
from io import BytesIO

from openpyxl import load_workbook

from odoo import _, fields, models
from odoo.exceptions import UserError

COL_ACCOUNT = 0
COL_DATE = 1
COL_DESC = 5
COL_DEBIT = 8
COL_CREDIT = 9


def _parse_number(val):
    if val is None:
        return 0.0
    if isinstance(val, int | float):
        return float(val)
    s = str(val).strip().replace("\u00a0", "")
    if not s:
        return 0.0
    if "," in s and "." in s:
        if s.rfind(",") > s.rfind("."):
            s = s.replace(".", "").replace(",", ".")
        else:
            s = s.replace(",", "")
    elif "," in s and "." not in s:
        s = s.replace(",", ".")
    try:
        return float(s)
    except Exception as exc:
        raise UserError(_("Numeric parse error for value: %s") % val) from exc


def _parse_date_cell(val):
    if not val:
        return None
    if isinstance(val, dt_date | datetime):
        return val.date() if isinstance(val, datetime) else val
    s = str(val).strip()
    for fmt in ("%d.%m.%Y", "%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y"):
        try:
            return datetime.strptime(s, fmt).date()
        except Exception:
            continue
    raise UserError(_("Could not parse date value: %s") % val)


class AccountAccrualUploadWizard(models.TransientModel):
    _name = "account.accrual.upload.wizard"
    _description = "Accrual Upload Wizard"

    data_file = fields.Binary(string=".xlsx File", required=True)
    filename = fields.Char()

    journal_id = fields.Many2one(
        "account.journal",
        default=lambda self: self.env["account.journal"].browse(93),
        required=True,
    )
    currency_id = fields.Many2one(
        "res.currency",
        default=lambda self: self.env["res.currency"].browse(31),
        required=True,
        readonly=True,
    )

    def _load_active_sheet(self):
        if not self.data_file:
            raise UserError(_("Please upload an .xlsx file."))

        try:
            decoded = base64.b64decode(self.data_file)
            wb = load_workbook(
                filename=BytesIO(decoded),
                data_only=True,
                read_only=True,
            )
        except Exception as exc:
            raise UserError(_("Could not read the .xlsx file: %s") % exc) from exc

        ws = wb.active
        if ws.max_row < 2:
            raise UserError(_("The uploaded file is empty."))
        return ws

    def _extract_move_date(self, row, current_date, row_index):
        if len(row) > COL_DATE and row[COL_DATE]:
            parsed = _parse_date_cell(row[COL_DATE])
            if parsed:
                if current_date is None:
                    return parsed
                if current_date != parsed:
                    msg = _(
                        "Row %(idx)s has a different date (%(row_date)s) than "
                        "previous rows (%(move_date)s). All rows must share "
                        "the same date."
                    ) % {
                        "idx": row_index,
                        "row_date": parsed,
                        "move_date": current_date,
                    }
                    raise UserError(msg)
        return current_date

    def _build_line_from_row(self, row, row_index, missing_accounts):
        account_code = (
            str(row[COL_ACCOUNT]).strip()
            if len(row) > COL_ACCOUNT and row[COL_ACCOUNT] is not None
            else ""
        )
        if not account_code:
            return None

        account = self.env["account.account"].search(
            [("code", "=", account_code)], limit=1
        )
        if not account:
            missing_accounts.add(account_code)
            return None

        desc = (
            str(row[COL_DESC]).strip()
            if len(row) > COL_DESC and row[COL_DESC] is not None
            else ""
        ) or "/"

        debit = _parse_number(row[COL_DEBIT] if len(row) > COL_DEBIT else 0.0)
        credit = _parse_number(row[COL_CREDIT] if len(row) > COL_CREDIT else 0.0)

        if not debit and not credit:
            return None
        if debit and credit:
            raise UserError(
                _("Row %s has both Debit and Credit. Please fix the file.") % row_index
            )

        return {
            "name": desc,
            "account_id": account.id,
            "debit": debit,
            "credit": credit,
            "currency_id": self.currency_id.id,
            "amount_currency": debit - credit,
        }

    def action_import(self):
        self.ensure_one()
        ws = self._load_active_sheet()

        move_date = None
        total_debit = 0.0
        total_credit = 0.0
        missing_accounts = set()
        line_vals = []

        for idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
            if idx == 1 or not row:
                continue

            move_date = self._extract_move_date(row, move_date, idx)

            line_dict = self._build_line_from_row(row, idx, missing_accounts)
            if not line_dict:
                continue

            total_debit += line_dict["debit"]
            total_credit += line_dict["credit"]
            line_vals.append((0, 0, line_dict))

        if missing_accounts:
            raise UserError(
                _("Missing account(s) for codes: %s")
                % ", ".join(sorted(missing_accounts))
            )

        if not line_vals:
            raise UserError(_("No valid lines found."))

        if move_date is None:
            move_date = fields.Date.context_today(self)

        if abs(total_debit - total_credit) > 0.01:
            msg = _(
                "The entry is not balanced.\n"
                "Total Debit: %(debit).2f\n"
                "Total Credit: %(credit).2f"
            ) % {"debit": total_debit, "credit": total_credit}
            raise UserError(msg)

        move_vals = {
            "move_type": "entry",
            "date": move_date,
            "journal_id": self.journal_id.id,
            "line_ids": line_vals,
            "ref": self.filename or _("Accrual Upload"),
        }
        move = self.env["account.move"].create(move_vals)
        move.action_post()

        action = self.env["ir.actions.act_window"]._for_xml_id(
            "account.action_move_journal_line"
        )
        action.update(
            {
                "view_mode": "form",
                "res_id": move.id,
                "views": [(False, "form")],
                "target": "current",
            }
        )
        return action
